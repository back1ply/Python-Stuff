import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import pandas as pd
import time
import random
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import json
import os

import ssl
import os

# Create unverified SSL context
ssl._create_default_https_context = ssl._create_unverified_context

try:
    import chromedriver_autoinstaller
except ImportError:
    chromedriver_autoinstaller = None

class USPhoneBookScraper:
    def __init__(self, input_csv="test_phonenumbers.csv", output_excel="phone_lookup_results.xlsx"):
        self.input_csv = input_csv
        self.output_excel = output_excel
        self.driver = None
        self.results = []
        self.checkpoint_file = "scraper_checkpoint.json"
        self.processed_numbers = set()
        self.load_checkpoint()
        
    def load_checkpoint(self):
        """Load previously processed numbers from checkpoint file"""
        if os.path.exists(self.checkpoint_file):
            try:
                with open(self.checkpoint_file, 'r') as f:
                    data = json.load(f)
                    self.processed_numbers = set(data.get('processed_numbers', []))
                    self.results = data.get('results', [])
                print(f"Loaded checkpoint: {len(self.processed_numbers)} numbers already processed")
            except Exception as e:
                print(f"Error loading checkpoint: {e}")
    
    def save_checkpoint(self):
        """Save current progress to checkpoint file"""
        try:
            checkpoint_data = {
                'processed_numbers': list(self.processed_numbers),
                'results': self.results
            }
            with open(self.checkpoint_file, 'w') as f:
                json.dump(checkpoint_data, f)
        except Exception as e:
            print(f"Error saving checkpoint: {e}")
    
    def setup_driver(self):
        """Setup undetected Chrome driver following working pattern"""
        try:
            print("Setting up undetected Chrome driver...")
            
            # Configure options
            options = uc.ChromeOptions()
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_argument('--window-size=1280,720')
            options.add_argument('--disable-extensions')
            options.add_argument('--disable-infobars')
            
            # Set page load strategy
            options.page_load_strategy = 'eager'
            
            # Set user agent
            options.add_argument('--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36')
            
            # Explicitly set non-headless mode (optional, as it's the default)
            # options.add_argument('--no-headless')  # Uncomment if needed for debugging
            
            # Create driver with minimal parameters
            self.driver = uc.Chrome(
                options=options,
                version_main=135,  # Ensure this matches your Chrome version
                use_subprocess=True  # Helps avoid some initialization issues
            )
            
            # Set timeouts
            self.driver.set_page_load_timeout(30)
            
            # Hide webdriver property
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            print("Driver successfully initialized!")
            return True
            
        except Exception as e:
            print(f"Error setting up driver: {e}")
            import traceback
            traceback.print_exc()  # Print full stack trace for debugging
            return False

    def handle_cloudflare_checkbox(self):
        """Attempt to handle the Cloudflare checkbox challenge"""
        try:
            # Wait for the Cloudflare challenge iframe to be present
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "iframe[src*='challenges.cloudflare.com'], iframe[title*='Widget'], iframe#challenge-form"))
            )
            
            # Find the iframe
            iframe = None
            possible_selectors = [
                "iframe[src*='challenges.cloudflare.com']",
                "iframe[title*='Widget']",
                "iframe#challenge-form",
                "iframe[name*='cf-challenge']"
            ]
            
            for selector in possible_selectors:
                try:
                    iframe = self.driver.find_element(By.CSS_SELECTOR, selector)
                    if iframe:
                        break
                except:
                    continue
            
            if not iframe:
                print("Could not find Cloudflare challenge iframe")
                return False
            
            # Switch to the iframe
            self.driver.switch_to.frame(iframe)
            
            # Wait for the checkbox to be present
            checkbox = WebDriverWait(self.driver, 15).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "input[type='checkbox'], .checkbox, .cb-lb"))
            )
            
            # Scroll to checkbox
            self.driver.execute_script("arguments[0].scrollIntoView(true);", checkbox)
            time.sleep(random.uniform(0.5, 1))
            
            # Click the checkbox using JavaScript for better reliability
            self.driver.execute_script("arguments[0].click();", checkbox)
            print("Clicked Cloudflare checkbox")
            
            # Wait for the challenge to be completed
            time.sleep(random.uniform(2, 3))
            
            # Switch back to main content
            self.driver.switch_to.default_content()
            
            return True
            
        except Exception as e:
            print(f"Error handling Cloudflare checkbox: {e}")
            try:
                self.driver.switch_to.default_content()
            except:
                pass
            return False
    

    def read_phone_numbers(self):
        """Read phone numbers from the CSV file"""
        try:
            df = pd.read_csv(self.input_csv)
            # Assuming the column is named 'Phone Number' based on your CSV
            phone_numbers = df['Phone Number'].tolist()
            return [str(phone).strip() for phone in phone_numbers if pd.notna(phone)]
        except Exception as e:
            print(f"Error reading CSV: {e}")
            return []
    
    def human_type(self, element, text):
        """Type text in a human-like manner with random delays"""
        for char in text:
            element.send_keys(char)
            time.sleep(random.uniform(0.05, 0.15))
    
    def format_phone_number(self, phone):
        """Format phone number for searching - no formatting, just clean digits"""
        # Keep only numeric characters
        phone = re.sub(r'\D', '', phone)
        return phone  # Return as-is without formatting
    
    def search_phone_number(self, phone_number):
        """Search for a phone number on USPhoneBook"""
        try:
            # Navigate to USPhoneBook
            print(f"Navigating to USPhoneBook for {phone_number}...")
            self.driver.get("https://www.usphonebook.com/")
            
            # Random delay to appear human
            time.sleep(random.uniform(2, 4))
            
            # Wait for the page to load completely
            WebDriverWait(self.driver, 15).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            
            # Check if we need to handle initial Cloudflare challenge
            if not self.wait_for_cloudflare(timeout=30):
                print("Initial page load failed Cloudflare check")
                return False
            
            # Wait for the phone input field
            phone_input = WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.ID, "focusPhone"))
            )
            
            # Human-like interaction with random mouse movements
            self.driver.execute_script("""
                var evt = new MouseEvent('mousemove', {
                    clientX: Math.random() * window.innerWidth,
                    clientY: Math.random() * window.innerHeight
                });
                document.dispatchEvent(evt);
            """)
            time.sleep(random.uniform(0.2, 0.4))
            
            phone_input.click()
            time.sleep(random.uniform(0.5, 1))
            
            # Clear and type phone number
            phone_input.clear()
            self.human_type(phone_input, phone_number)
            
            # Wait before submitting
            time.sleep(random.uniform(0.5, 1.5))
            
            # Submit the form
            search_success = False
            methods = [
                ("Click button", lambda: self.driver.find_element(By.CSS_SELECTOR, "#search1Form button").click()),
                ("Submit form", lambda: self.driver.execute_script("document.getElementById('search1Form').submit();")),
                ("Press Enter", lambda: phone_input.send_keys(Keys.RETURN))
            ]
            
            for method_name, method_func in methods:
                try:
                    print(f"Trying to search using: {method_name}")
                    method_func()
                    search_success = True
                    break
                except Exception as e:
                    print(f"{method_name} failed: {e}")
                    continue
            
            if not search_success:
                print(f"Failed to submit search for {phone_number}")
                return False
            
            # Wait for the page to start loading
            time.sleep(random.uniform(1, 2))
            
            # Wait for Cloudflare check to complete with extended timeout
            print("Waiting for search results page to load...")
            if not self.wait_for_cloudflare(timeout=90):
                print(f"Failed to pass Cloudflare check for {phone_number}")
                return False
            
            # Additional wait for page to fully load
            time.sleep(random.uniform(2, 3))
            
            # Verify we're on a results page
            try:
                # Look for either search results or no results message
                WebDriverWait(self.driver, 10).until(
                    lambda d: d.find_elements(By.CSS_SELECTOR, ".ls_contacts-btn, .no-results, .phase2-section, #search-results-holder")
                )
                print(f"Successfully loaded search results for {phone_number}")
                return True
            except TimeoutException:
                print(f"Search results page not found for {phone_number}")
                return False
                
        except TimeoutException:
            print(f"Timeout while searching for {phone_number}")
            return False
        except Exception as e:
            print(f"Error searching for {phone_number}: {e}")
            return False
    
    def wait_for_cloudflare(self, timeout=60):
        """Wait for Cloudflare check to complete or confirm USPhoneBook page is loaded"""
        print("Checking for Cloudflare or USPhoneBook page load...")
        start_time = time.time()
        checkbox_attempts = 0
        max_checkbox_attempts = 3

        while time.time() - start_time < timeout:
            try:
                # First, check for USPhoneBook page elements
                try:
                    # Check for homepage, results, or details page elements
                    elements = self.driver.find_elements(By.CSS_SELECTOR, "#focusPhone, .ls_contacts-btn, .no-results, .phase2-section, #search-results-holder")
                    if elements:
                        print("USPhoneBook page loaded successfully, no Cloudflare challenge detected!")
                        return True
                except:
                    pass  # No USPhoneBook elements found yet

                # Check if we're on a Cloudflare page
                title = self.driver.title.lower()
                url = self.driver.current_url.lower()
                page_source = self.driver.page_source.lower()

                # Debug: Log page title and source snippet
                print(f"Debug - Page title: {title}")
                print(f"Debug - Page source snippet: {page_source[:500]}")

                # Specific Cloudflare indicators
                cloudflare_indicators = [
                    'just a moment',
                    'checking your browser',
                    'ray id',
                    'ddos protection by cloudflare',
                    'verify you are human'
                ]

                is_cloudflare = any(indicator in title or indicator in url or indicator in page_source for indicator in cloudflare_indicators)

                if not is_cloudflare:
                    # Double-check for USPhoneBook elements
                    try:
                        self.driver.find_elements(By.CSS_SELECTOR, "#focusPhone, .ls_contacts-btn, .no-results, .phase2-section, #search-results-holder")
                        print("Confirmed USPhoneBook page loaded, no Cloudflare detected!")
                        return True
                    except:
                        print("No recognizable USPhoneBook elements found, continuing to check...")
                        time.sleep(random.uniform(1, 2))
                        continue

                # Check for Cloudflare challenge checkbox (more specific)
                if 'challenges.cloudflare.com' in page_source or 'cf-chl-btn' in page_source:
                    if checkbox_attempts < max_checkbox_attempts:
                        print(f"Attempting to solve Cloudflare checkbox challenge (attempt {checkbox_attempts + 1})")
                        if self.handle_cloudflare_checkbox():
                            checkbox_attempts += 1
                            time.sleep(random.uniform(2, 4))
                            continue
                        else:
                            print("Failed to handle Cloudflare checkbox")
                    else:
                        print("Max checkbox attempts reached")
                
                # Wait before checking again
                time.sleep(random.uniform(1, 2))

            except Exception as e:
                print(f"Error while checking for Cloudflare: {e}")
                time.sleep(random.uniform(1, 2))

        print(f"Cloudflare check timed out after {timeout} seconds")
        return False

    def click_get_details(self):
        """Click the 'Get Details' button if present"""
        try:
            # Look for the Get Details button with a longer wait
            get_details_buttons = []
            
            # Try multiple selectors for the Get Details button
            selectors = [
                "a.ls_contacts-btn.newcolor",
                "a.ls_contacts-btn",
                "a[href*='/']:contains('Get Details')",
                ".ls_contacts-btn"
            ]
            
            for selector in selectors:
                try:
                    buttons = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for button in buttons:
                        if 'get details' in button.text.lower():
                            get_details_buttons.append(button)
                except:
                    continue
            
            if not get_details_buttons:
                print("No Get Details button found")
                return False
            
            # Click the first Get Details button found
            button = get_details_buttons[0]
            
            # Scroll to the button
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
            time.sleep(random.uniform(0.5, 1))
            
            # Click the button
            print("Clicking Get Details button...")
            button.click()
            
            # Wait for the details page to load
            print("Waiting for details page to load...")
            try:
                WebDriverWait(self.driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".phase2-section"))
                )
                print("Details page loaded successfully!")
            except TimeoutException:
                print("Details page not found after clicking Get Details")
                # Check for Cloudflare only if details page didn't load
                print("Checking for Cloudflare challenge...")
                if not self.wait_for_cloudflare(timeout=30):
                    print("Failed to load details page or pass Cloudflare check")
                    return False
            
            # Confirm details page is loaded
            try:
                self.driver.find_element(By.CSS_SELECTOR, ".phase2-section")
                print("Confirmed details page is loaded")
                return True
            except:
                print("Could not confirm details page load")
                return False
                
        except Exception as e:
            print(f"Error clicking Get Details: {e}")
            return False

    def extract_details(self, phone_number):
        """Extract all information from the details page"""
        details = {
            'Phone Number': phone_number,
            'Name': '',
            'Age': '',
            'Current Address': '',
            'Previous Addresses': [],
            'Phone Numbers': [],
            'Relatives': [],
            'Associates': [],
            'Email': ''
        }
        
        try:
            # Extract name
            try:
                name_element = self.driver.find_element(By.CSS_SELECTOR, ".phase2-section h3 span")
                details['Name'] = name_element.text.strip()
            except:
                print("Name not found")
            
            # Extract age
            try:
                age_element = self.driver.find_element(By.CSS_SELECTOR, ".ls_contacts__age")
                details['Age'] = age_element.text.strip()
            except:
                print("Age not found")
            
            # Extract current address
            try:
                current_address = self.driver.find_element(By.CSS_SELECTOR, "a[href*='/address/'] .ls_contacts__text")
                details['Current Address'] = current_address.text.strip()
            except:
                print("Current address not found")
            
            # Extract previous addresses using xpath for more reliable selection
            try:
                prev_addresses_section = self.driver.find_element(By.XPATH, "//h3[contains(text(), 'Previous Addresses')]")
                # Get all list items in the next ul after the heading
                addresses = self.driver.find_elements(By.XPATH, "//h3[contains(text(), 'Previous Addresses')]/following-sibling::ul//li//a")
                details['Previous Addresses'] = [addr.text.strip() for addr in addresses if addr.text.strip()]
            except:
                print("Previous addresses not found")
            
            # Extract previous phone numbers
            try:
                phone_section = self.driver.find_element(By.XPATH, "//h3[contains(text(), 'Previous Phone Numbers')]")
                phone_div = phone_section.find_element(By.XPATH, "following-sibling::div[1]")
                phone_text = phone_div.text.strip()
                if phone_text and "There is currently no phone numbers data available" not in phone_text:
                    details['Phone Numbers'] = [phone_text]
            except:
                print("Previous phone numbers not found")
            
            # Extract relatives
            try:
                relatives_elements = self.driver.find_elements(By.CSS_SELECTOR, ".relative-card span[itemprop='name']")
                details['Relatives'] = [rel.text.strip() for rel in relatives_elements if rel.text.strip()]
            except:
                print("Relatives not found")
            
            # Extract associates
            try:
                # Look for Associates section
                associates_section = self.driver.find_element(By.XPATH, "//h3[contains(text(), 'Associates')]")
                # Get all associate cards in the section-relative div that follows
                associates_elements = self.driver.find_elements(By.XPATH, "//h3[contains(text(), 'Associates')]/following-sibling::div[contains(@class, 'section-relative')]//span")
                details['Associates'] = [assoc.text.strip() for assoc in associates_elements if assoc.text.strip() and 'more available' not in assoc.text.lower()]
            except:
                print("Associates not found")
            
            # Extract email
            try:
                email_section = self.driver.find_element(By.XPATH, "//h3[contains(text(), 'Email')]")
                email_div = email_section.find_element(By.XPATH, "following-sibling::div[1]")
                email_text = email_div.text.strip()
                if email_text and "There is currently no email data available" not in email_text:
                    details['Email'] = email_text
            except:
                print("Email not found")
            
            print(f"Successfully extracted details for {phone_number}")
                
        except Exception as e:
            print(f"Error extracting details for {phone_number}: {e}")
            
        return details
    
    def save_to_excel(self):
        """Save all results to an Excel file"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Phone Lookup Results"
            
            # Define headers
            headers = [
                'Phone Number', 'Name', 'Age', 'Current Address', 'Previous Addresses',
                'Previous Phone Numbers', 'Relatives', 'Associates', 'Email'
            ]
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            for row, result in enumerate(self.results, 2):
                ws.cell(row=row, column=1, value=result['Phone Number'])
                ws.cell(row=row, column=2, value=result['Name'])
                ws.cell(row=row, column=3, value=result['Age'])
                ws.cell(row=row, column=4, value=result['Current Address'])
                ws.cell(row=row, column=5, value='\n'.join(result['Previous Addresses']))
                ws.cell(row=row, column=6, value='\n'.join(result['Phone Numbers']))
                ws.cell(row=row, column=7, value='\n'.join(result['Relatives']))
                ws.cell(row=row, column=8, value='\n'.join(result['Associates']))
                ws.cell(row=row, column=9, value=result['Email'])
            
            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, 15), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            wb.save(self.output_excel)
            print(f"Results saved to {self.output_excel}")
            
        except Exception as e:
            print(f"Error saving to Excel: {e}")
    
    def run(self):
        """Main method to run the scraper"""
        try:
            # Setup driver
            print("Setting up browser...")
            if not self.setup_driver():
                print("Failed to setup driver. Exiting...")
                return
            
            # Read phone numbers
            print("Reading phone numbers from CSV...")
            phone_numbers = self.read_phone_numbers()
            
            if not phone_numbers:
                print("No phone numbers found in CSV")
                return
            
            print(f"Found {len(phone_numbers)} phone numbers to process")
            
            # Process each phone number
            for i, phone in enumerate(phone_numbers, 1):
                # Skip already processed numbers
                if phone in self.processed_numbers:
                    print(f"Skipping {phone} - already processed")
                    continue
                
                print(f"\nProcessing {i}/{len(phone_numbers)}: {phone}")
                
                # Format phone number (no formatting, just clean)
                formatted_phone = self.format_phone_number(phone)
                
                # Search for the phone number
                if self.search_phone_number(formatted_phone):
                    # Click Get Details button
                    if self.click_get_details():
                        # Extract details
                        details = self.extract_details(phone)
                        self.results.append(details)
                        self.processed_numbers.add(phone)
                        print(f"Successfully extracted details for {phone}")
                    else:
                        print(f"Could not find details page for {phone}")
                        # Still add a basic record
                        self.results.append({
                            'Phone Number': phone,
                            'Name': 'Details not available',
                            'Age': '',
                            'Current Address': '',
                            'Previous Addresses': [],
                            'Phone Numbers': [],
                            'Relatives': [],
                            'Associates': [],
                            'Email': ''
                        })
                        self.processed_numbers.add(phone)
                else:
                    print(f"Could not search for {phone}")
                
                # Save checkpoint after each number
                self.save_checkpoint()
                
                # Save to Excel after each phone number
                print(f"Updating Excel file after {phone}...")
                self.save_to_excel()
                
                # Add delay between requests to avoid detection (limited to max 5 seconds)
                delay = random.uniform(2, 5)
                print(f"Waiting {delay:.1f} seconds before next request...")
                time.sleep(delay)
            
            # Final save to Excel (redundant but ensures everything is saved)
            print("\nSaving final results to Excel...")
            self.save_to_excel()
            
        except Exception as e:
            print(f"Error in main execution: {e}")
            # Save checkpoint before exiting
            self.save_checkpoint()
            import traceback
            traceback.print_exc()
        
        finally:
            # Clean up
            if self.driver:
                self.driver.quit()
                print("\nBrowser closed")

# Usage
if __name__ == "__main__":
    # Create scraper instance
    scraper = USPhoneBookScraper(input_csv=r"C:\Users\Elya Pollak\Desktop\Programs\enrich data\test_phonenumbers.csv",
        output_excel="phone_lookup_results.xlsx"
    )
    
    # Run the scraper
    scraper.run()